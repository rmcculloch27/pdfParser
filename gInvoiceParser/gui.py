import tkinter as tk
from tkinter import filedialog, messagebox
import os
import pandas as pd
import subprocess
import platform
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from .parser import SuperHeroFlex

# Setup logging
logging.basicConfig(
    level=logging.DEBUG,
    filename="gInvoiceParser_debug.log",
    filemode="w",
    format="%(asctime)s - %(levelname)s - %(message)s"
)

REQUIRED_FIELDS = [
    "InvoiceType", "Invoice#", "Month", "Description", "Amount($)"
]

HIGHLIGHT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

class InvoiceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Invoice PDF Extractor")
        self.file_paths = []
        self.output_path = None

        self.label = tk.Label(root, text="Select PDF files to process")
        self.label.pack(pady=10)

        self.select_button = tk.Button(root, text="Select PDFs", command=self.select_files)
        self.select_button.pack(pady=5)

        self.run_button = tk.Button(root, text="Extract & Export to Excel", command=self.process_files)
        self.run_button.pack(pady=5)

        self.view_button = tk.Button(root, text="View File", command=self.open_file, state="disabled")
        self.view_button.pack(pady=5)

        self.status = tk.Label(root, text="No files selected.", fg="blue")
        self.status.pack(pady=10)

    def select_files(self):
        self.file_paths = filedialog.askopenfilenames(
            title="Select PDF Files",
            filetypes=[("PDF files", "*.pdf")]
        )
        self.status.config(text=f"{len(self.file_paths)} file(s) selected.")
        logging.debug(f"Selected files: {self.file_paths}")

    def process_files(self):
        if not self.file_paths:
            messagebox.showwarning("No Files", "Please select PDF files first.")
            return

        try:
            parser = SuperHeroFlex(self.file_paths)
            combined_df = parser.extract_all()  
        except Exception as e:
            messagebox.showerror("Processing Error", f"Unexpected error: {e}")
            logging.exception("Error during PDF parsing:")
            return

        if combined_df.empty:
            self.status.config(text="No valid data extracted.")
            return

        path_result = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save Excel File"
        )

        if isinstance(path_result, tuple):
            logging.warning(f"asksaveasfilename returned a tuple: {path_result}")
            output_path = path_result[0]
        else:
            output_path = path_result

        logging.debug(f"Resolved output_path: {output_path} (type: {type(output_path)})")

        if output_path:
            try:
                combined_df = combined_df.astype(str).fillna("N/A")

                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    combined_df.to_excel(writer, sheet_name="Invoice Data", index=False)

                self.output_path = output_path
                self.view_button.config(state="normal")
                self.status.config(text=f"✅ Exported to {os.path.basename(output_path)}")

            except Exception as export_error:
                messagebox.showerror("Export Error", f"Failed to write Excel file:\n{export_error}")
                logging.exception("Error during Excel export:")
        else:
            self.status.config(text="Export canceled.")
            logging.info("Export canceled by user.")

    def open_file(self):
        if self.output_path:
            try:
                if platform.system() == "Darwin":
                    subprocess.call(["open", self.output_path])
                elif platform.system() == "Windows":
                    os.startfile(self.output_path)
                elif platform.system() == "Linux":
                    subprocess.call(["xdg-open", self.output_path])
            except Exception as e:
                logging.exception("Error opening the exported file:")

def main():
    root = tk.Tk()
    app = InvoiceApp(root)
    root.mainloop()
